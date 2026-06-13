# 주식 트렌드 데이터를 4시트 구조의 인사이트 엑셀로 빌드하는 모듈

import json, os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from datetime import datetime

C_HEADER_BG = "1A1F36"
C_HEADER_FG = "FFFFFF"
C_ACCENT1   = "2563EB"
C_ACCENT2   = "E53935"
C_LIGHT_BLUE = "EFF6FF"
C_BORDER    = "E5E9F2"

SECTOR_KEYWORDS = [
    "이차전지","반도체","바이오","AI반도체","전기차","방산","원전",
    "로봇","수소","태양광","게임","엔터","핀테크","클라우드","조선",
    "철강","제약","의료기기","인공지능","NVIDIA","Tesla","AMD"
]

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def apply_header_row(ws, row, headers, widths, bg=C_HEADER_BG, fg=C_HEADER_FG):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color=fg, size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        set_col_width(ws, i, w)

def dfont(bold=False, color="1A1F36", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def load_history():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_trend_history.xlsx")
    if os.path.exists(path):
        df = pd.read_excel(path)
        df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.strftime("%Y-%m-%d")
        df = df.dropna(subset=["날짜"])
        return df
    return pd.DataFrame()


def run():
    df = load_history()
    wb = Workbook()
    wb.remove(wb.active)
    summary = []

    # 시트1: 일별_신호
    ws1 = wb.create_sheet("📅 일별_신호")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"
    ws1.merge_cells("A1:K1")
    t = ws1["A1"]
    t.value = "📊 일별 신호 누적 데이터"
    t.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=13)
    t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32
    h1 = ["날짜","순위","종목/키워드","총점","당일점수","3일점수","7일점수","주가_당일(%)","주가_7일(%)","신호내용","소스"]
    w1 = [13,7,18,8,9,8,8,13,13,45,25]
    apply_header_row(ws1, 2, h1, w1)
    ws1.row_dimensions[2].height = 22

    if not df.empty:
        for r_idx, row in df.iterrows():
            excel_row = r_idx + 3
            try:
                score = float(row.get("총점", 0) or 0)
            except Exception:
                score = 0
            bg = "FEF2F2" if score >= 7 else ("EFF6FF" if score >= 4 else "F8F9FC")
            row_data = [
                str(row.get("날짜",""))[:10], row.get("순위",""), row.get("종목/키워드",""),
                row.get("총점",""), row.get("당일점수",""), row.get("3일점수",""), row.get("7일점수",""),
                row.get("주가_당일(%)",""), row.get("주가_7일(%)",""), row.get("신호내용",""), row.get("소스",""),
            ]
            for c_idx, val in enumerate(row_data, 1):
                c = ws1.cell(row=excel_row, column=c_idx, value=val)
                c.font = dfont(bold=(c_idx==3))
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="center" if c_idx not in [3,10,11] else "left", vertical="center")
                if c_idx in [8,9]:
                    try:
                        fval = float(val)
                        if fval > 0:
                            c.font = Font(name="Arial", bold=True, color=C_ACCENT2, size=10)
                        elif fval < 0:
                            c.font = Font(name="Arial", bold=True, color=C_ACCENT1, size=10)
                    except Exception:
                        pass
            ws1.row_dimensions[excel_row].height = 18
        ws1.auto_filter.ref = f"A2:K{len(df)+2}"

    # 시트2: 종목_추적
    ws2 = wb.create_sheet("🎯 종목_추적")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:I1")
    t2 = ws2["A1"]
    t2.value = "🎯 종목별 신호 추적 - 연속 감지 및 누적 분석"
    t2.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=13)
    t2.fill = PatternFill("solid", fgColor="7C3AED")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32
    h2 = ["종목/키워드","총감지횟수","최고점수","평균점수","연속감지일","최초감지일","최근감지일","평균주가당일(%)","평균주가7일(%)"]
    w2 = [18,12,10,10,12,13,13,16,16]
    apply_header_row(ws2, 2, h2, w2, bg="7C3AED")
    ws2.row_dimensions[2].height = 22

    if not df.empty:
        for name, g in df.groupby("종목/키워드"):
            g = g.sort_values("날짜")
            dates = sorted(g["날짜"].dropna().astype(str).unique())
            max_streak = streak = 1
            for i in range(1, len(dates)):
                try:
                    d1 = pd.to_datetime(dates[i-1])
                    d2 = pd.to_datetime(dates[i])
                    if (d2 - d1).days <= 1:
                        streak += 1
                        max_streak = max(max_streak, streak)
                    else:
                        streak = 1
                except Exception:
                    streak = 1
            avg_d1 = pd.to_numeric(g["주가_당일(%)"], errors="coerce").mean()
            avg_d7 = pd.to_numeric(g["주가_7일(%)"], errors="coerce").mean()
            summary.append({
                "종목": name, "횟수": len(g),
                "최고점수": pd.to_numeric(g["총점"], errors="coerce").max(),
                "평균점수": round(pd.to_numeric(g["총점"], errors="coerce").mean(), 1),
                "연속감지일": max_streak,
                "최초": dates[0][:10] if dates else "-",
                "최근": dates[-1][:10] if dates else "-",
                "평균당일": round(avg_d1, 2) if pd.notna(avg_d1) else None,
                "평균7일": round(avg_d7, 2) if pd.notna(avg_d7) else None,
            })
        summary.sort(key=lambda x: (x["연속감지일"], x["최고점수"] or 0), reverse=True)
        for r_idx, s in enumerate(summary, 3):
            vals = [s["종목"],s["횟수"],s["최고점수"],s["평균점수"],s["연속감지일"],s["최초"],s["최근"],s["평균당일"],s["평균7일"]]
            bg = "FEF2F2" if s["연속감지일"] >= 3 else ("FFFBEB" if s["연속감지일"] >= 2 else "F8F9FC")
            for c_idx, val in enumerate(vals, 1):
                c = ws2.cell(row=r_idx, column=c_idx, value=val)
                c.font = dfont(bold=(c_idx in [1,5]))
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="left" if c_idx==1 else "center", vertical="center")
                if c_idx in [8,9]:
                    try:
                        if val and float(val) > 0:
                            c.font = Font(name="Arial", bold=True, color=C_ACCENT2, size=10)
                        elif val and float(val) < 0:
                            c.font = Font(name="Arial", bold=True, color=C_ACCENT1, size=10)
                    except Exception:
                        pass
            ws2.row_dimensions[r_idx].height = 18
        if summary:
            ws2.auto_filter.ref = "A2:I" + str(len(summary)+2)

    # 시트3: 섹터_트렌드
    ws3 = wb.create_sheet("📈 섹터_트렌드")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A3"
    ws3.merge_cells("A1:G1")
    t3 = ws3["A1"]
    t3.value = "📈 섹터 키워드 트렌드 분석"
    t3.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=13)
    t3.fill = PatternFill("solid", fgColor="16A34A")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32
    h3 = ["섹터/키워드","총감지횟수","최고점수","평균점수","최근감지일","감지날짜목록","비고"]
    w3 = [18,12,10,10,13,35,20]
    apply_header_row(ws3, 2, h3, w3, bg="16A34A")
    ws3.row_dimensions[2].height = 22

    if not df.empty:
        sector_data = []
        for kw in SECTOR_KEYWORDS:
            g = df[df["종목/키워드"] == kw]
            if g.empty:
                sector_data.append({"kw":kw,"cnt":0,"max_s":0,"avg_s":0,"last":"-","dates":"감지 없음","note":""})
            else:
                dates = sorted(g["날짜"].dropna().astype(str).unique())
                sector_data.append({
                    "kw":kw, "cnt":len(g),
                    "max_s": pd.to_numeric(g["총점"], errors="coerce").max(),
                    "avg_s": round(pd.to_numeric(g["총점"], errors="coerce").mean(), 1),
                    "last": dates[-1][:10] if dates else "-",
                    "dates": ", ".join(d[:10] for d in dates[-5:]),
                    "note": "🔥 주목" if len(g) >= 3 else ("⚡ 관심" if len(g) >= 1 else "")
                })
        sector_data.sort(key=lambda x: x["cnt"], reverse=True)
        for r_idx, s in enumerate(sector_data, 3):
            vals = [s["kw"],s["cnt"],s["max_s"],s["avg_s"],s["last"],s["dates"],s["note"]]
            bg = "FEF2F2" if s["cnt"] >= 3 else ("FFFBEB" if s["cnt"] >= 1 else "F8F9FC")
            for c_idx, val in enumerate(vals, 1):
                c = ws3.cell(row=r_idx, column=c_idx, value=val)
                c.font = dfont(bold=(c_idx==1))
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="left" if c_idx in [1,6,7] else "center", vertical="center")
            ws3.row_dimensions[r_idx].height = 18

    # 시트4: 대시보드 (첫 탭)
    ws4 = wb.create_sheet("📊 대시보드", 0)
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:N1")
    mt = ws4["A1"]
    now_str = datetime.today().strftime("%Y-%m-%d %H:%M")
    mt.value = "📊 주식 트렌드 조기 감지 대시보드   |   최종 업데이트: " + now_str
    mt.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=14)
    mt.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    mt.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 36

    if not df.empty:
        kpi_values = [
            df["날짜"].nunique(),
            df["종목/키워드"].nunique(),
            len(df[pd.to_numeric(df["총점"], errors="coerce") >= 7]["종목/키워드"].unique()),
            len([s for s in summary if s["연속감지일"] >= 3]),
            len(df)
        ]
    else:
        kpi_values = [0,0,0,0,0]

    kpi_labels = ["총 분석일수","분석 종목수","강한신호 종목","3일 연속 감지","누적 데이터 수"]
    kpi_colors = [C_ACCENT1,"16A34A",C_ACCENT2,"7C3AED","F59E0B"]
    for i, (label, val, color) in enumerate(zip(kpi_labels, kpi_values, kpi_colors)):
        cs = i*3+1
        ce = cs+1
        ls = get_column_letter(cs)
        le = get_column_letter(ce)
        ws4.merge_cells(ls+"3:"+le+"4")
        ws4.merge_cells(ls+"5:"+le+"6")
        nc = ws4[ls+"3"]
        nc.value = val
        nc.font = Font(name="Arial", bold=True, color=color, size=22)
        nc.fill = PatternFill("solid", fgColor="FFFFFF")
        nc.alignment = Alignment(horizontal="center", vertical="center")
        nc.border = thin_border()
        lc = ws4[ls+"5"]
        lc.value = label
        lc.font = Font(name="Arial", color="6B7280", size=10)
        lc.fill = PatternFill("solid", fgColor="F8F9FC")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = thin_border()
        for r in [3,4,5,6]:
            for col in range(cs, ce+1):
                cell = ws4.cell(row=r, column=col)
                cell.border = thin_border()
                cell.fill = PatternFill("solid", fgColor="FFFFFF" if r in [3,4] else "F8F9FC")
            ws4.row_dimensions[r].height = 26
    for i in range(1, 16):
        ws4.column_dimensions[get_column_letter(i)].width = 10

    ws4.row_dimensions[8].height = 10
    ws4.merge_cells("A9:G9")
    ins = ws4["A9"]
    ins.value = "💡 주요 인사이트"
    ins.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=11)
    ins.fill = PatternFill("solid", fgColor=C_ACCENT1)
    ins.alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[9].height = 24

    insights = []
    if not df.empty and summary:
        for s in [s for s in summary if s["연속감지일"] >= 2][:3]:
            insights.append("🔥 " + str(s["종목"]) + "  →  " + str(s["연속감지일"]) + "일 연속 감지, 최고점수 " + str(s["최고점수"]) + "점")
        recent_date = df["날짜"].dropna().max()
        df_r = df[df["날짜"] == recent_date].copy()
        df_r["총점_num"] = pd.to_numeric(df_r["총점"], errors="coerce")
        for _, row in df_r.nlargest(3, "총점_num").iterrows():
            insights.append("⚡ 최신(" + str(recent_date)[:10] + ") 상위: " + str(row["종목/키워드"]) + "  점수 " + str(row["총점"]) + "점")
    else:
        insights = ["데이터가 쌓이면 인사이트가 자동으로 표시됩니다."]

    for i, txt in enumerate(insights[:6], 10):
        ws4.merge_cells("A"+str(i)+":G"+str(i))
        c = ws4["A"+str(i)]
        c.value = txt
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=C_LIGHT_BLUE if i%2==0 else "FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin_border()
        ws4.row_dimensions[i].height = 20

    if not df.empty and df["날짜"].nunique() >= 2:
        ws4.merge_cells("I9:N9")
        ch = ws4["I9"]
        ch.value = "📈 날짜별 강한 신호 종목 수"
        ch.font = Font(name="Arial", bold=True, color=C_HEADER_FG, size=11)
        ch.fill = PatternFill("solid", fgColor="7C3AED")
        ch.alignment = Alignment(horizontal="left", vertical="center")
        df_num = df.copy()
        df_num["총점_num"] = pd.to_numeric(df_num["총점"], errors="coerce")
        date_counts = df_num[df_num["총점_num"] >= 5].groupby("날짜")["종목/키워드"].count().reset_index()
        date_counts.columns = ["날짜","종목수"]
        csr = 10
        ws4.cell(row=csr, column=9, value="날짜")
        ws4.cell(row=csr, column=10, value="종목수")
        for r_i, (_, row) in enumerate(date_counts.iterrows(), csr+1):
            ws4.cell(row=r_i, column=9, value=str(row["날짜"])[:10])
            ws4.cell(row=r_i, column=10, value=int(row["종목수"]))
        chart = LineChart()
        chart.title = "날짜별 신호 종목 수"
        chart.style = 10
        chart.height = 8
        chart.width = 16
        data_ref = Reference(ws4, min_col=10, min_row=csr, max_row=csr+len(date_counts))
        chart.add_data(data_ref, titles_from_data=True)
        ws4.add_chart(chart, "I11")

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_trend_history.xlsx")
    wb.save(out_path)
    print("Excel 저장 완료 -> " + out_path)


if __name__ == "__main__":
    run()
