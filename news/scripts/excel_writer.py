# Excel watchlist 업데이트 모듈 (openpyxl 기반)

import sys
import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

BASE_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(BASE_DIR, "watchlist.xlsx")

# 색상 정의
COLOR_KR_HEADER = "1F4E79"   # 국내 헤더 (진한 파랑)
COLOR_US_HEADER = "7B2D00"   # 미국 헤더 (진한 갈색)
COLOR_SUBHEADER = "BDD7EE"   # 서브 헤더 (연한 파랑)
COLOR_ROW_ODD   = "F2F7FC"   # 홀수 행 배경
COLOR_ROW_EVEN  = "FFFFFF"   # 짝수 행 배경
COLOR_LOG_HDR   = "375623"   # 로그 시트 헤더 (초록)


def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, bg_color, font_color="FFFFFF", bold=True, size=11):
    cell.font      = Font(bold=bold, color=font_color, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = make_border()


def style_cell(cell, bg_color="FFFFFF", bold=False, wrap=True):
    cell.font      = Font(bold=bold, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border    = make_border()


def init_workbook():
    """새 workbook 생성 (시트1: 종목현황, 시트2: 뉴스전체)"""
    from config import KR_STOCKS, US_STOCKS

    wb = openpyxl.Workbook()

    # ── 시트1: 종목 현황 (스냅샷) ──────────────────────────
    ws1 = wb.active
    ws1.title = "종목현황"

    kr_headers = ["#", "종목명", "섹터", "최신 뉴스 제목", "출처", "뉴스 날짜", "투자 시사점", "링크"]
    us_headers = ["#", "티커", "종목명", "섹터", "최신 뉴스 제목", "출처", "뉴스 날짜", "투자 시사점", "링크"]

    row = 1

    # 국내 섹션 타이틀
    ws1.merge_cells(f"A{row}:H{row}")
    c = ws1.cell(row=row, column=1, value="🇰🇷 국내 종목 (KRX)")
    style_header(c, COLOR_KR_HEADER, size=12)
    row += 1

    # 국내 컬럼 헤더
    for col, h in enumerate(kr_headers, 1):
        c = ws1.cell(row=row, column=col, value=h)
        style_header(c, COLOR_SUBHEADER, font_color="1F4E79")
    row += 1

    for i, s in enumerate(KR_STOCKS):
        bg = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
        vals = [i+1, s["name"], s["sector"], "", "", "", "", ""]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=v)
            style_cell(c, bg)
        row += 1

    row += 1  # 빈 줄

    # 미국 섹션 타이틀
    ws1.merge_cells(f"A{row}:I{row}")
    c = ws1.cell(row=row, column=1, value="🇺🇸 미국 종목 (NYSE/NASDAQ)")
    style_header(c, COLOR_US_HEADER, size=12)
    row += 1

    # 미국 컬럼 헤더
    for col, h in enumerate(us_headers, 1):
        c = ws1.cell(row=row, column=col, value=h)
        style_header(c, "F4CCCC", font_color="7B2D00")
    row += 1

    for i, s in enumerate(US_STOCKS):
        bg = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
        vals = [i+1, s["ticker"], s["name"], s["sector"], "", "", "", "", ""]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=v)
            style_cell(c, bg)
        row += 1

    # 컬럼 너비 설정
    col_widths_kr = [4, 16, 16, 50, 14, 12, 40, 30]
    for i, w in enumerate(col_widths_kr, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── 시트2: 뉴스 전체 (누적 로그) ──────────────────────
    ws2 = wb.create_sheet("뉴스전체")
    log_headers = ["수집일", "시장", "종목명", "티커", "섹터", "뉴스 제목", "출처", "뉴스 날짜", "투자 시사점", "링크"]
    for col, h in enumerate(log_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        style_header(c, COLOR_LOG_HDR)

    col_widths_log = [12, 8, 16, 8, 18, 55, 14, 12, 40, 35]
    for i, w in enumerate(col_widths_log, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = "A2"
    ws1.freeze_panes = "A3"

    wb.save(XLSX_PATH)
    print(f"Excel 초기 생성 완료: {XLSX_PATH}")


def update_snapshot(stock_name: str, ticker: str, market: str,
                    title: str, source: str, news_date: str,
                    implication: str, url: str):
    """시트1 해당 종목 행 최신 뉴스로 덮어쓰기"""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["종목현황"]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == stock_name:
                r = cell.row
                if market == "KRX":
                    ws.cell(r, 4).value = title
                    ws.cell(r, 5).value = source
                    ws.cell(r, 6).value = news_date
                    ws.cell(r, 7).value = implication
                    ws.cell(r, 8).value = url
                else:
                    ws.cell(r, 5).value = title
                    ws.cell(r, 6).value = source
                    ws.cell(r, 7).value = news_date
                    ws.cell(r, 8).value = implication
                    ws.cell(r, 9).value = url
                break

    wb.save(XLSX_PATH)


def append_log(collected_date: str, market: str, stock_name: str,
               ticker: str, sector: str, title: str, source: str,
               news_date: str, implication: str, url: str):
    """시트2 뉴스 전체 로그에 행 추가"""
    wb  = openpyxl.load_workbook(XLSX_PATH)
    ws  = wb["뉴스전체"]
    row = ws.max_row + 1
    bg  = COLOR_ROW_ODD if row % 2 == 0 else COLOR_ROW_EVEN

    vals = [collected_date, market, stock_name, ticker, sector,
            title, source, news_date, implication, url]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        style_cell(c, bg)

    wb.save(XLSX_PATH)


def bulk_update(all_news: dict, collected_date: str):
    """
    수집된 전체 뉴스를 받아 시트1 스냅샷 + 시트2 로그 동시 업데이트
    all_news = {
      "stock_name": [{"title","source","date","implication","url","ticker","market","sector"}, ...]
    }
    """
    from config import KR_STOCKS, US_STOCKS
    # Excel 파일 없으면 자동 초기화
    if not os.path.exists(XLSX_PATH):
        print("[안내] watchlist.xlsx 없음 → 자동 생성")
        init_workbook()

    stock_meta = {}
    for s in KR_STOCKS:
        stock_meta[s["name"]] = {"ticker": "", "market": "KRX", "sector": s["sector"]}
    for s in US_STOCKS:
        stock_meta[s["name"]] = {"ticker": s["ticker"], "market": "US", "sector": s["sector"]}

    for stock_name, items in all_news.items():
        if not items:
            continue
        meta = stock_meta.get(stock_name, {})

        # 스냅샷: 가장 최신 뉴스 1건만
        latest = items[0]
        update_snapshot(
            stock_name  = stock_name,
            ticker      = meta.get("ticker", ""),
            market      = meta.get("market", ""),
            title       = latest.get("title", ""),
            source      = latest.get("source", ""),
            news_date   = latest.get("date", ""),
            implication = latest.get("implication", ""),
            url         = latest.get("url", ""),
        )

        # 로그: 모든 뉴스 추가
        for item in items:
            append_log(
                collected_date = collected_date,
                market         = meta.get("market", ""),
                stock_name     = stock_name,
                ticker         = meta.get("ticker", ""),
                sector         = meta.get("sector", ""),
                title          = item.get("title", ""),
                source         = item.get("source", ""),
                news_date      = item.get("date", ""),
                implication    = item.get("implication", ""),
                url            = item.get("url", ""),
            )

    print(f"Excel 업데이트 완료 ({collected_date})")


# CLI 실행: python3 excel_writer.py init | python3 excel_writer.py update <news.json> <date>
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: excel_writer.py init | excel_writer.py update <news.json> <date>")
        sys.exit(1)

    cmd = sys.argv[1]
    if cmd == "init":
        init_workbook()
    elif cmd == "update" and len(sys.argv) == 4:
        with open(sys.argv[2], "r", encoding="utf-8") as f:
            data = json.load(f)
        bulk_update(data, sys.argv[3])
