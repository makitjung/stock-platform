# 종목별 뉴스 Excel 파일 누적 저장 모듈

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import KR_STOCKS, US_STOCKS, BASE_DIR

KR_NAMES   = {s["name"] for s in KR_STOCKS}
KR_DIR     = os.path.join(BASE_DIR, "국내")
US_DIR     = os.path.join(BASE_DIR, "미국")

COLOR_KR   = "1F4E79"   # 국내 헤더 (진한 파랑)
COLOR_US   = "7B2D00"   # 미국 헤더 (진한 갈색)
COLOR_KR_ROW = "EBF3FB" # 국내 짝수 행
COLOR_US_ROW = "FBF3EB" # 미국 짝수 행
COLOR_WHITE  = "FFFFFF"


def _get_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"],
            check=True
        )
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    return openpyxl, Font, PatternFill, Alignment, Border, Side


def _stock_dir(stock_name: str) -> str | None:
    for s in KR_STOCKS:
        if s["name"] == stock_name:
            return os.path.join(KR_DIR, s["folder"])
    for s in US_STOCKS:
        if s["name"] == stock_name:
            return os.path.join(US_DIR, s["folder"])
    return None


def _make_border(Side):
    t = Side(style="thin", color="CCCCCC")
    return __import__("openpyxl").styles.Border(left=t, right=t, top=t, bottom=t)


def _init_sheet(ws, stock_name: str, Font, PatternFill, Alignment, Side):
    """새 시트 헤더 초기화."""
    is_kr = stock_name in KR_NAMES
    hdr_color = COLOR_KR if is_kr else COLOR_US

    headers = ["수집일", "뉴스날짜", "제목 (클릭하면 기사 이동)", "출처"]
    widths  = [12, 12, 68, 10]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _make_border(Side)

    col_letters = ["A", "B", "C", "D"]
    for letter, w in zip(col_letters, widths):
        ws.column_dimensions[letter].width = w

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def save_stock_news_excel(all_news: dict, today: str) -> None:
    """
    수집된 뉴스를 종목별 Excel 파일에 누적 저장.
    - 중복 제거: 같은 날짜 + 제목 앞 60자가 동일하면 스킵
    - 국내: 국내/{folder}/{종목명}_뉴스.xlsx
    - 미국: 미국/{folder}/{종목명}_뉴스.xlsx
    """
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()

    total_added = 0

    for stock_name, items in all_news.items():
        if not items:
            continue

        stock_dir = _stock_dir(stock_name)
        if not stock_dir:
            continue
        os.makedirs(stock_dir, exist_ok=True)

        is_kr      = stock_name in KR_NAMES
        hdr_color  = COLOR_KR if is_kr else COLOR_US
        row_color  = COLOR_KR_ROW if is_kr else COLOR_US_ROW
        xlsx_path  = os.path.join(stock_dir, f"{stock_name}_뉴스.xlsx")

        # 기존 파일 로드 or 신규 생성
        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            # 기존 항목 키 수집 (중복 체크)
            existing = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = str(row[1]) if row[1] else ""
                t = str(row[2])[:60] if row[2] else ""
                if d or t:
                    existing.add((d, t))
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "뉴스"
            existing = set()
            _init_sheet(ws, stock_name, Font, PatternFill, Alignment, Side)

        border = _make_border(Side)
        added = 0

        for item in items:
            title  = item.get("title", "").strip()
            date   = item.get("date", "")
            url    = item.get("url", "")
            source = item.get("source", "")

            key = (date, title[:60])
            if key in existing:
                continue
            existing.add(key)

            rn  = ws.max_row + 1
            bg  = row_color if rn % 2 == 0 else COLOR_WHITE

            def cell(col, val):
                c = ws.cell(row=rn, column=col, value=val)
                c.fill   = PatternFill("solid", fgColor=bg)
                c.border = border
                return c

            # 수집일
            c1 = cell(1, today)
            c1.font      = Font(size=10, color="555555")
            c1.alignment = Alignment(horizontal="center", vertical="center")

            # 뉴스날짜
            c2 = cell(2, date)
            c2.font      = Font(size=10, bold=True)
            c2.alignment = Alignment(horizontal="center", vertical="center")

            # 제목 (하이퍼링크)
            c3 = cell(3, title)
            if url:
                c3.hyperlink = url
                c3.font = Font(size=10, color="1155CC", underline="single")
            else:
                c3.font = Font(size=10)
            c3.alignment = Alignment(vertical="center", wrap_text=True)

            # 출처
            c4 = cell(4, source.upper())
            c4.font      = Font(size=9, bold=True,
                                color="1F4E79" if source == "naver" else "7B2D00")
            c4.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[rn].height = 36
            added += 1

        if added:
            wb.save(xlsx_path)
            total_added += added
            print(f"  [{stock_name}] +{added}건")
        else:
            print(f"  [{stock_name}] 신규 없음 (중복 스킵)")

    print(f"[종목별 Excel 저장 완료] 총 {total_added}건 추가")
